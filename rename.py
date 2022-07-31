# 用类封装
class HandExecl:
    def __init__(self, input_file_name, relation_sheet_name, output_file_name, ToBe_modified_shtname):
        self.input_file_name = input_file_name   #
        self.relation_sheet_name = relation_sheet_name
        self.output_file_name = output_file_name
        self.ToBe_modified_shtname = ToBe_modified_shtname


        self.obj_relation_wb = self.load_wb(self.input_file_name)
        self.dict_tuple = self.create_relation_dict()
        self.out_result_wb = self.wb_exists()

    def get_sht_maxrow(self, ws, reference_col_num=1):  # 获取表的最大行数；
        list_row_num = []
        for row in ws.rows:
            cell = row[reference_col_num - 1]
            if cell.value:
                list_row_num.append(cell.row)
        sht_maxrow = max(list_row_num)
        return sht_maxrow

    def load_wb(self,file_name):  # 以只读方式读入待处理的表；
        p = Path.cwd()
        wb = openpyxl.load_workbook(p /file_name, data_only=True, read_only=True)
        return wb

    def wb_exists(self):  # 判断输出结果表是否存在，无论存在与否都重新创建；
        p = Path.cwd()
        path = Path(p / self.output_file_name)
        x = path.exists()  # 路径是否存在
        y = path.is_file()  # 是否是个文件
        exists = x and y
        if exists:
            path.unlink()  # 结果工作簿存在则删除工作簿
        out_result_wb = openpyxl.Workbook(write_only=True)  # 以只写模式创建
        out_result_wb.create_sheet(title=self.ToBe_modified_shtname)
        return out_result_wb

    def create_result_sht(self, shtname):  # 创建输出工作表结构；
        p = Path.cwd()
        result_sht = self.out_result_wb.create_sheet(title=shtname)
        return result_sht

    def modify_customer_name(self, storage_wb_shtname):
        ROW_NUM_MAX = 8  # 操作单元格范围
        START_ROW_NUM = 2  # 操作单元格范围
        collect_set = set()
        storage_wb_sheet = self.obj_relation_wb[storage_wb_shtname]  # 生成listwbname(工作表)对象；
        count_sht_num = 0

        for rownum in range(START_ROW_NUM, ROW_NUM_MAX + 1):
            if storage_wb_sheet.cell(rownum, 1).value != None:
                count_sht_num += 1  # 统计总的要处理表的总数；

                quote_wb_name = storage_wb_sheet.cell(rownum, 1).value  # 要处理工作簿名
                quote_sheet_name = storage_wb_sheet.cell(rownum, 2).value  # 要处理的表名
                quote_start_rownum = storage_wb_sheet.cell(rownum, 3).value  # 起始行号
                quote_sheet_specify_colnum = storage_wb_sheet.cell(rownum, 4).value  # 要处理的指定列；
                #            quote_sheet_reference_col=storage_wb_sheet.cell(rownum,5).value
                #            Identifier=storage_wb_sheet.cell(rownum,6).value                    #标识码
                quote_wb = self.load_wb(quote_wb_name)  # 载入要处理的工作簿
                quote_sheet = quote_wb[quote_sheet_name]  # 要处理的表名称
                #quote_sheet_maxrow = storage_wb_sheet.cell(rownum, 5).value  # 获取每个表的最大行数
                result_sht = self.create_result_sht(quote_sheet_name + "#" + str(count_sht_num))  # 避免相同的表重名

                count_rownum = 0
                for row in quote_sheet.rows:
                    list1 = []  # 收集每行数据
                    for cell in row:
                        list1.append(cell.value)
                    count_rownum += 1
                    if count_rownum >= quote_start_rownum:  # 修改值操作；
                        logo = False
                        cell_value = row[quote_sheet_specify_colnum - 1].value  # 获取客户名称所在列单元格的值；
                        if cell_value != None:
                            for dict in self.dict_tuple :
                                list_keys = list(dict.keys())
                                list_values = list(dict.values())
                                generator_enumerate = enumerate(list_values)
                                if str(cell_value) not in dict:
                                    return_value = self.get_customer_group_name(cell_value,
                                                                           generator_enumerate)  # 调用函数，修改客户名称;
                                    if return_value != None:
                                        list1[quote_sheet_specify_colnum - 1] = list_keys[
                                            return_value]  # 修改客户为客户群组中的名称；
                                        logo = True
                                else:
                                    logo = True
                                    break
                            if logo == False:
                                collect_set.add(cell_value)  # 不在字典中则单独收集为没有对应客户群组的待处理客户；

                    result_sht.append(list1)  # 写入每行数据
                quote_wb.close()  # 处理完毕后关闭工作簿
        collect_list = list(collect_set)  # 汇总所有表的待处理客户
        self.out_result_wb[self.ToBe_modified_shtname].append(collect_list)  # 将所有待处理客户写入到输出表中；


    def get_customer_group_name(self,cell_value,generator_enumerate):   #检验客户简称中是否再字典中存在对应的键；
        logo=False
        for index,elem in generator_enumerate:
            for  value in  elem:
                if str(cell_value).upper()==str(value).upper():  #改成分块查找，relation表前半部分为主要出货大客户，后边部分随着出货量递减排列；
                    logo=True
                    return index                                  #程序也同时修改
                    break
            if logo==True:
                break
        if logo==False:
            return None


    def create_relation_dict(self):
        # 创建映射字典#
        relation_sheet = self.obj_relation_wb[self.relation_sheet_name]
        relation_sheet_maxrow = self.get_sht_maxrow(relation_sheet, reference_col_num=1)
        START_ROW_NUM = 3
        START_COL_NUM = 2
        RELATION_SHEET_MAXCOL = 15  # 列范围
        dict_important_customer = {}  # 最主要客户字典(优先查找)；
        dict_secondary_customer = {}  # 次要客户字典(此先查找)；
        dict_small_customer = {}  # 小客户字典(最后查找)；

        for rownum in range(START_ROW_NUM, relation_sheet_maxrow + 1):
            cell_key = relation_sheet.cell(rownum, 1).value  # 字典键；
            degree_of_importance = relation_sheet.cell(rownum, 19).value  # 获取客户的重要性值；

            for colnum in range(START_COL_NUM, RELATION_SHEET_MAXCOL + 1):
                cell_value = relation_sheet.cell(rownum, colnum).value  # 获取键对应值，一对多的映射字典；
                if cell_value != None:
                    if degree_of_importance == 1:
                        dict_important_customer.setdefault(cell_key, []).append(cell_value)  # 写入重要客户的映射字典；
                    elif degree_of_importance == 2:
                        dict_secondary_customer.setdefault(cell_key, []).append(cell_value)  # 写入次要客户的映射字典；
                    else:
                        dict_small_customer.setdefault(cell_key, []).append(cell_value)  # 写入小客户的映射字典；

        dict_tuple = (dict_important_customer, dict_secondary_customer, dict_small_customer)  # 将三个字典放入到元组中；
        return dict_tuple


if __name__ == "__main__":
    import gc
    import sys
    import openpyxl
    from pathlib import Path

    storage_wb_shtname = "listwbname"

    try:
        hc=HandExecl("客户群组对应表.xlsx","relation","OutputResult.xlsx","ToBeModified")
    except PermissionError:
        print("待处理的Excel文件处于打开状态,请关闭后再运行")
    else:
        try:
            hc.modify_customer_name(storage_wb_shtname)
        except TypeError:
            print(f"请确认[{hc.input_file_name}]工作簿中的listwbname表,标识行或列的单元格值不为空。")
        except FileNotFoundError:
            print(f"请确认[{hc.input_file_name}]工作簿中的listwbname表,A列显示的表均放在目录下面")
      #  except:
      #     print("(2)其他未知错误，程序已经退出")
        else:
            print("程序运行正常，无异常报出。")
            hc.out_result_wb.save(hc.output_file_name) #无异常的情况下，保存输出工作簿；
    finally:
        hc.obj_relation_wb.close()    #关闭只读模式下打开的映射表(客户群组映射表)；
        try:
            sys.exit(0)            #其他情况下强行退出Python程序；
        except SystemExit:
            del hc.obj_relation_wb   #应对在jupyter notebook中，表obj_relation_wb被Python一直占用，无法保存的情况；
            gc.collect()




